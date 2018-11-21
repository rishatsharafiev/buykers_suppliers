import decimal
import zipfile

import openpyxl
from conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError, transaction

from ..models import Good, Picture, Task
import logging


class TaskParser:
    """Task parser"""

    def run(self, obj):
        """Run"""
        # First file
        file_path = settings.MEDIA_ROOT / str(obj.first_file)
        workbook = openpyxl.load_workbook(filename=file_path, read_only=True)
        self.save_motos(obj, workbook)

        # Second file
        file_path = settings.MEDIA_ROOT / str(obj.second_file)
        workbook = openpyxl.load_workbook(filename=file_path, read_only=True)
        self.get_desctiption(workbook)

        # Zip file
        zip_file_path = settings.MEDIA_ROOT / str(obj.zip_file)
        self.image_processing(zip_file_path)

    def save_motos(self, task, workbook):
        """Bicycles sheet parser"""
        worksheet = workbook['Мотоэкипировка']

        genders = list(zip(*Good.GENDER_CHOICES))
        genders_indexes = genders[0]
        genders_titles = genders[1]

        try:
            task.status = Task.STATUS_CHOICE_PROGRESS
            task.save()

            # remove objects
            Good.objects.all().delete()

            # write objects
            goods = []

            for row in worksheet.iter_rows(row_offset=3):
                code = str(row[0].value).strip()
                vendor_code = row[1].value
                nomenclature = row[2].value
                nomenclature_group = row[3].value
                brand = row[4].value
                if row[4].value == 1:
                    brand = '100%'

                status = str(row[6].value)
                if status.strip() == 'Ожидается':
                    continue

                try:
                    if str(row[7].value).strip() == 'в наличии':
                        count = 1
                    else:
                        count = int(row[7].value)
                except TypeError:
                    continue
                try:
                    wholesale_price = decimal.Decimal(row[8].value)
                    retail_price = decimal.Decimal(row[9].value)
                except (TypeError, decimal.InvalidOperation):
                    continue
                try:
                    gender_index = genders_titles.index(row[14].value)
                    gender = genders_indexes[gender_index]
                except ValueError:
                    gender = Good.GENDER_EMPTY_CHOICE

                good = Good(
                    code=code,
                    vendor_code=vendor_code,
                    nomenclature=nomenclature,
                    nomenclature_group=nomenclature_group,
                    brand=brand,
                    count=count,
                    wholesale_price=wholesale_price,
                    retail_price=retail_price,
                    gender=gender,
                    task=task
                )
                goods.append(good)

            Good.objects.bulk_create(goods, 100)
            task.status = Task.STATUS_CHOICE_DONE
        except IntegrityError:
            task.status = Task.STATUS_CHOICE_ERROR
        finally:
            task.save()

    def get_desctiption(self, workbook):
        """Get goods description"""
        worksheet = workbook['balances']

        with transaction.atomic():
            for row in worksheet.iter_rows(row_offset=1):
                try:
                    good = Good.objects.get(code=str(row[0].value).strip())
                    good.descriptions = row[7].value
                    good.save()
                except ObjectDoesNotExist as err:
                    logging.warning(str(err))

    def image_processing(self, path):
        """Image processing"""
        if not zipfile.is_zipfile(path):
            return

        z = zipfile.ZipFile(path, 'r')
        names = z.namelist()
        for name in names:
            with transaction.atomic():

                # Name example: A04398_1.jpeg
                code = name[:name.index('.')].split('_')[0].strip()
                try:
                    good = Good.objects.get(code=code)
                except ObjectDoesNotExist:
                    continue

                picture, _ = Picture.objects.get_or_create(
                    name=name,
                    good=good
                )
                picture.save()
